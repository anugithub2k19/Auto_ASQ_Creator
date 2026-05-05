from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


MANDATORY_COLUMNS = [
    "Tenant",
    "Project",
    "Title",
    "Module",
    "Severity",
    "Priority",
    "Issue Category",
    "Sub Category",
    "Owner",
    "Authorized Closer",
    "Description",
]

OPTIONAL_COLUMNS = [
    "Run",
    "Attachment Path",
]


@dataclass(frozen=True)
class TicketRow:
    row_index_1based: int
    tenant: str
    project: str
    title: str
    module: str
    severity: str
    priority: str
    issue_category: str
    sub_category: str
    owner: str
    authorized_closer: str
    description: str
    run: bool = True
    attachment_path: Optional[str] = None


def _norm_header(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _to_bool_run(v: Any) -> bool:
    if v is None:
        return True
    s = str(v).strip().lower()
    if s in {"y", "yes", "true", "1"}:
        return True
    if s in {"n", "no", "false", "0"}:
        return False
    return True


def load_tickets(xlsx_path: str | Path, sheet_name: str) -> list[TicketRow]:
    p = Path(xlsx_path)
    wb = load_workbook(filename=str(p), data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [_norm_header(v) for v in header_row]
    index = {h: i for i, h in enumerate(headers) if h}

    missing = [c for c in MANDATORY_COLUMNS if c not in index]
    if missing:
        raise ValueError(
            "Missing mandatory columns: "
            + ", ".join(missing)
            + ". Expected at least: "
            + ", ".join(MANDATORY_COLUMNS)
        )

    out: list[TicketRow] = []
    for excel_row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None:
            continue

        def get(col: str) -> str:
            val = row[index[col]] if index[col] < len(row) else None
            return "" if val is None else str(val).strip()

        run = _to_bool_run(row[index["Run"]]) if "Run" in index else True
        attachment_path = (
            get("Attachment Path") if "Attachment Path" in index else ""
        ) or None

        if not any((get(c) for c in MANDATORY_COLUMNS)) and attachment_path is None:
            continue

        out.append(
            TicketRow(
                row_index_1based=excel_row_idx,
                tenant=get("Tenant"),
                project=get("Project"),
                title=get("Title"),
                module=get("Module"),
                severity=get("Severity"),
                priority=get("Priority"),
                issue_category=get("Issue Category"),
                sub_category=get("Sub Category"),
                owner=get("Owner"),
                authorized_closer=get("Authorized Closer"),
                description=get("Description"),
                run=run,
                attachment_path=attachment_path,
            )
        )
    return out


def write_results(
    *,
    template_xlsx: str | Path,
    sheet_name: str,
    results_xlsx: str | Path,
    per_row_results: dict[int, tuple[str, str]],
) -> None:
    src = Path(template_xlsx)
    dst = Path(results_xlsx)
    dst.parent.mkdir(parents=True, exist_ok=True)

    wb: Workbook = load_workbook(filename=str(src))
    ws = wb[sheet_name]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [_norm_header(v) for v in header_row]

    def ensure_col(name: str) -> int:
        if name in headers:
            return headers.index(name) + 1
        headers.append(name)
        ws.cell(row=1, column=len(headers)).value = name
        return len(headers)

    status_col = ensure_col("Status")
    error_col = ensure_col("Error")

    for row_idx, (status, err) in per_row_results.items():
        ws.cell(row=row_idx, column=status_col).value = status
        ws.cell(row=row_idx, column=error_col).value = err

    wb.save(str(dst))

