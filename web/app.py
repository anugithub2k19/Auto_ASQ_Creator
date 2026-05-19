from flask import Flask, render_template, request, jsonify, send_file
import os
import sys
from pathlib import Path
from datetime import datetime
import threading
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))
from qraie_ticket_bot.bot import QRaieBot
from qraie_ticket_bot.config import load_config
from qraie_ticket_bot.excel_io import load_tickets

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['UPLOAD_FOLDER'] = Path(__file__).parent.parent / 'uploads'
app.config['LOG_FOLDER'] = Path(__file__).parent.parent / 'logs'

app.config['UPLOAD_FOLDER'].mkdir(parents=True, exist_ok=True)
app.config['LOG_FOLDER'].mkdir(parents=True, exist_ok=True)

status_lock = threading.Lock()
automation_status = {
    'running': False,
    'progress': 0,
    'message': '',
    'results': [],
    'start_time': '',
    'end_time': ''
}

def validate_excel(file_path):
    """Validate if the uploaded Excel file has the correct structure."""
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
        
        required_columns = [
            'Tenant', 'Project', 'Title', 'Module', 'Severity',
            'Priority', 'Issue Category', 'Sub Category', 'Owner',
            'Authorized Closer', 'Description'
        ]
        
        missing = [col for col in required_columns if col not in headers]
        if missing:
            return False, f"Missing columns: {', '.join(missing)}"
        
        if ws.max_row < 2:
            return False, "Excel file has no data rows"
            
        return True, "Valid Excel file"
    except Exception as e:
        return False, str(e)

def run_automation(excel_path, config_path):
    """Run the ticket creation automation in a background thread."""
    global automation_status
    try:
        with status_lock:
            automation_status['running'] = True
            automation_status['progress'] = 0
            automation_status['message'] = 'Loading configuration...'
            automation_status['start_time'] = datetime.now().strftime('%H:%M:%S')
        
        cfg = load_config(config_path)
        tickets = load_tickets(excel_path, sheet_name=cfg.run.sheet_name)
        
        total_tickets = len(tickets)
        
        def on_progress(completed, total, status):
            with status_lock:
                pct = 10 + int((completed / total) * 90) if total > 0 else 10
                automation_status['progress'] = min(pct, 99)
                automation_status['message'] = f'Processing ticket {completed}/{total}...'
        
        with status_lock:
            automation_status['progress'] = 10
            automation_status['message'] = f'Found {total_tickets} tickets to create...'
        
        bot = QRaieBot(cfg)
        results = bot.run(tickets, progress_callback=on_progress)
        
        # Generate log file
        log_entries = []
        success_count = 0
        fail_count = 0
        
        for row_idx, result in sorted(results.items()):
            status = "PASS" if result.status == "CREATED" else "FAIL"
            if status == "PASS":
                success_count += 1
            else:
                fail_count += 1
                
            log_entries.append({
                'row': row_idx,
                'status': status,
                'message': result.error if result.error else 'Ticket created successfully'
            })
        
        with status_lock:
            automation_status['results'] = log_entries
            automation_status['success_count'] = success_count
            automation_status['fail_count'] = fail_count
            automation_status['progress'] = 100
            automation_status['message'] = f'Completed: {success_count} passed, {fail_count} failed'
        
        # Save log file
        log_path = app.config['LOG_FOLDER'] / f'automation_log_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt'
        with open(log_path, 'w') as f:
            f.write(f"Automation Log - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 60 + "\n\n")
            f.write(f"Total Rows Processed: {len(log_entries)}\n")
            f.write(f"Passed: {success_count}\n")
            f.write(f"Failed: {fail_count}\n\n")
            f.write("-" * 60 + "\n")
            f.write(f"{'Row':<8} {'Status':<10} {'Message'}\n")
            f.write("-" * 60 + "\n")
            for entry in log_entries:
                f.write(f"{entry['row']:<8} {entry['status']:<10} {entry['message']}\n")
        
        with status_lock:
            automation_status['log_file'] = str(log_path)
        
    except Exception as e:
        with status_lock:
            automation_status['running'] = False
            automation_status['progress'] = 0
            automation_status['message'] = f'Error: {str(e)}'
    finally:
        with status_lock:
            automation_status['running'] = False
            automation_status['end_time'] = datetime.now().strftime('%H:%M:%S')

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'error': 'Invalid file type. Please upload an Excel file (.xlsx or .xls)'}), 400
    
    # Save uploaded file
    file_path = app.config['UPLOAD_FOLDER'] / f'uploaded_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    file.save(str(file_path))
    
    # Validate Excel file
    is_valid, message = validate_excel(file_path)
    if not is_valid:
        os.remove(file_path)
        return jsonify({'success': False, 'error': message}), 400
    
    return jsonify({
        'success': True,
        'message': message,
        'file_path': str(file_path)
    })

@app.route('/migrate', methods=['POST'])
def migrate():
    global automation_status
    
    with status_lock:
        if automation_status['running']:
            return jsonify({'success': False, 'error': 'Automation is already running'}), 400
        
        data = request.get_json()
        excel_path = data.get('file_path')
        
        if not excel_path or not os.path.exists(excel_path):
            return jsonify({'success': False, 'error': 'Invalid file path'}), 400
        
        # Reset status
        automation_status = {
            'running': False,
            'progress': 0,
            'message': 'Starting...',
            'results': [],
            'start_time': '',
            'end_time': ''
        }
    
    config_path = str(Path(__file__).parent.parent / 'config.yaml')
    
    # Run automation in background thread
    thread = threading.Thread(target=run_automation, args=(excel_path, config_path))
    thread.start()
    
    return jsonify({'success': True, 'message': 'Automation started'})

@app.route('/status')
def status():
    with status_lock:
        return jsonify(automation_status.copy())

@app.route('/download-log')
def download_log():
    log_files = list(app.config['LOG_FOLDER'].glob('*.txt'))
    if not log_files:
        return jsonify({'success': False, 'error': 'No log files available'}), 404
    
    # Get the most recent log file
    latest_log = max(log_files, key=os.path.getctime)
    return send_file(str(latest_log), as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True, port=5000)
